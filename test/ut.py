"""Pytest unit tests for city_tier_stats module."""

from __future__ import annotations

import csv
import io
import math
import textwrap
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook
from openpyxl import load_workbook as lw

from city_tier_stats import (
    TIER_LABELS,
    TIER_ORDER,
    build_location_details,
    calculate_tier_stats,
    city_match_keys,
    clean_csv_header,
    decode_csv_content,
    extract_city_name,
    is_empty_value,
    load_city_tier_mapping,
    match_city_tier,
    match_location_tier,
    normalize_city_name,
    read_csv_location_column,
    read_location_column,
    read_xlsx_location_column,
    resolve_config_path,
    sniff_csv_dialect,
    write_location_details,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(tmp_path: Path, rows: list[list]) -> Path:
    """Write rows to an xlsx file and return its path."""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


def _make_csv(tmp_path: Path, content: str, filename: str = "test.csv") -> Path:
    """Write CSV content (utf-8-sig encoded) and return its path."""
    path = tmp_path / filename
    path.write_bytes(content.encode("utf-8-sig"))
    return path


def _make_yaml(tmp_path: Path, data: dict) -> Path:
    """Write a YAML config file and return its path."""
    path = tmp_path / "tiers.yaml"
    path.write_text(yaml.dump(data, allow_unicode=True), encoding="utf-8")
    return path


SIMPLE_MAPPING: dict[str, str] = {
    "北京": "first",
    "上海": "first",
    "成都": "new_first",
    "武汉": "new_first",
    "厦门": "second",
}


# ---------------------------------------------------------------------------
# is_empty_value
# ---------------------------------------------------------------------------

class TestIsEmptyValue:
    def test_none_is_empty(self):
        assert is_empty_value(None) is True

    def test_nan_is_empty(self):
        assert is_empty_value(float("nan")) is True

    def test_zero_not_empty(self):
        assert is_empty_value(0) is False

    def test_empty_string_not_empty(self):
        assert is_empty_value("") is False

    def test_normal_string_not_empty(self):
        assert is_empty_value("北京") is False

    def test_integer_not_empty(self):
        assert is_empty_value(42) is False


# ---------------------------------------------------------------------------
# extract_city_name
# ---------------------------------------------------------------------------

class TestExtractCityName:
    def test_standard_dash_format(self):
        assert extract_city_name("广东省-深圳-南山区") == "深圳"

    def test_single_value_returns_itself(self):
        assert extract_city_name("重庆") == "重庆"

    def test_none_returns_empty(self):
        assert extract_city_name(None) == ""

    def test_nan_returns_empty(self):
        assert extract_city_name(float("nan")) == ""

    def test_empty_string_returns_empty(self):
        assert extract_city_name("") == ""

    def test_fullwidth_dash_normalized(self):
        # － (U+FF0D fullwidth hyphen-minus)
        assert extract_city_name("广东省－广州－天河区") == "广州"

    def test_em_dash_normalized(self):
        # — (U+2014 em dash)
        assert extract_city_name("广东省—东莞—南城") == "东莞"

    def test_en_dash_normalized(self):
        # – (U+2013 en dash)
        assert extract_city_name("广东省–佛山–禅城") == "佛山"

    def test_strips_whitespace(self):
        assert extract_city_name("  北京-  朝阳  -某街道  ") == "朝阳"

    def test_two_parts(self):
        assert extract_city_name("北京-朝阳") == "朝阳"


# ---------------------------------------------------------------------------
# normalize_city_name
# ---------------------------------------------------------------------------

class TestNormalizeCityName:
    def test_strips_shi_suffix(self):
        assert normalize_city_name("北京市") == "北京"

    def test_no_shi_suffix_unchanged(self):
        assert normalize_city_name("北京") == "北京"

    def test_none_returns_empty(self):
        assert normalize_city_name(None) == ""

    def test_strips_whitespace(self):
        assert normalize_city_name("  上海市  ") == "上海"

    def test_nan_returns_empty(self):
        assert normalize_city_name(float("nan")) == ""


# ---------------------------------------------------------------------------
# city_match_keys
# ---------------------------------------------------------------------------

class TestCityMatchKeys:
    def test_plain_city(self):
        assert city_match_keys("北京") == {"北京"}

    def test_city_with_shi(self):
        # normalize strips 市 first, then builds keys from "北京"
        assert city_match_keys("北京市") == {"北京"}

    def test_diqu_suffix(self):
        keys = city_match_keys("喀什地区")
        assert "喀什地区" in keys
        assert "喀什" in keys

    def test_meng_suffix(self):
        keys = city_match_keys("兴安盟")
        assert "兴安盟" in keys
        assert "兴安" in keys

    def test_zizhizhou_suffix(self):
        keys = city_match_keys("黔东南苗族侗族自治州")
        assert "黔东南苗族侗族自治州" in keys
        assert "黔东南苗族侗族" in keys

    def test_empty_returns_empty_set(self):
        assert city_match_keys("") == set()

    def test_none_returns_empty_set(self):
        assert city_match_keys(None) == set()


# ---------------------------------------------------------------------------
# match_city_tier
# ---------------------------------------------------------------------------

class TestMatchCityTier:
    def test_known_city_returns_tier(self):
        assert match_city_tier("北京", SIMPLE_MAPPING) == "first"

    def test_unknown_city_returns_other(self):
        assert match_city_tier("火星市", SIMPLE_MAPPING) == "other"

    def test_empty_city_returns_other(self):
        assert match_city_tier("", SIMPLE_MAPPING) == "other"

    def test_none_city_returns_other(self):
        assert match_city_tier(None, SIMPLE_MAPPING) == "other"

    def test_city_with_shi_matches(self):
        # mapping stores "北京" but input is "北京市" — normalize handles it
        assert match_city_tier("北京市", SIMPLE_MAPPING) == "first"


# ---------------------------------------------------------------------------
# match_location_tier
# ---------------------------------------------------------------------------

class TestMatchLocationTier:
    def test_compound_location(self):
        assert match_location_tier("广东省-深圳-南山区", {"深圳": "first"}) == "first"

    def test_unknown_location(self):
        assert match_location_tier("火星省-火星市-某区", SIMPLE_MAPPING) == "other"

    def test_none_location(self):
        assert match_location_tier(None, SIMPLE_MAPPING) == "other"


# ---------------------------------------------------------------------------
# calculate_tier_stats
# ---------------------------------------------------------------------------

class TestCalculateTierStats:
    def test_order_matches_tier_order(self):
        locations = ["北京", "上海", "成都", "火星"]
        mapping = {"北京": "first", "上海": "first", "成都": "new_first"}
        stats = calculate_tier_stats(locations, mapping)
        assert [row["tier"] for row in stats] == list(TIER_ORDER)

    def test_counts_correct(self):
        locations = ["北京", "上海", "成都", "火星"]
        mapping = {"北京": "first", "上海": "first", "成都": "new_first"}
        stats = calculate_tier_stats(locations, mapping)
        tier_map = {row["tier"]: row for row in stats}
        assert tier_map["first"]["count"] == 2
        assert tier_map["new_first"]["count"] == 1
        assert tier_map["other"]["count"] == 1

    def test_percentages_sum_to_one(self):
        locations = ["北京", "成都", "火星"]
        mapping = {"北京": "first", "成都": "new_first"}
        stats = calculate_tier_stats(locations, mapping)
        total_pct = sum(float(row["percentage"]) for row in stats)
        assert math.isclose(total_pct, 1.0)

    def test_empty_locations(self):
        stats = calculate_tier_stats([], SIMPLE_MAPPING)
        for row in stats:
            assert row["count"] == 0
            assert row["percentage"] == 0

    def test_labels_present(self):
        stats = calculate_tier_stats(["北京"], {"北京": "first"})
        tier_map = {row["tier"]: row for row in stats}
        for tier in TIER_ORDER:
            assert tier_map[tier]["label"] == TIER_LABELS[tier]


# ---------------------------------------------------------------------------
# build_location_details
# ---------------------------------------------------------------------------

class TestBuildLocationDetails:
    def test_returns_one_row_per_location(self):
        locations = ["广东省-深圳-南山区", "广东省-广州-天河区"]
        details = build_location_details(locations, {"深圳": "first", "广州": "new_first"})
        assert len(details) == 2

    def test_row_structure(self):
        details = build_location_details(["北京-朝阳"], {"北京": "first"})
        row = details[0]
        assert row["归属地"] == "北京-朝阳"
        assert row["城市"] == "朝阳"
        assert row["城市归一化"] == "朝阳"
        assert row["分层代码"] == "other"  # "朝阳" not in mapping
        assert row["分层"] == TIER_LABELS["other"]

    def test_known_city_row(self):
        details = build_location_details(["省-北京-区"], {"北京": "first"})
        row = details[0]
        assert row["分层代码"] == "first"
        assert row["分层"] == TIER_LABELS["first"]

    def test_none_location_row(self):
        details = build_location_details([None], SIMPLE_MAPPING)
        row = details[0]
        assert row["归属地"] is None
        assert row["城市"] == ""
        assert row["分层代码"] == "other"


# ---------------------------------------------------------------------------
# clean_csv_header
# ---------------------------------------------------------------------------

class TestCleanCsvHeader:
    def test_strips_whitespace(self):
        assert clean_csv_header("  归属地  ") == "归属地"

    def test_strips_bom(self):
        assert clean_csv_header("\ufeff归属地") == "归属地"

    def test_none_returns_empty(self):
        assert clean_csv_header(None) == ""

    def test_plain_header_unchanged(self):
        assert clean_csv_header("城市") == "城市"


# ---------------------------------------------------------------------------
# sniff_csv_dialect
# ---------------------------------------------------------------------------

class TestSniffCsvDialect:
    def test_comma_delimiter(self):
        dialect = sniff_csv_dialect("a,b,c\n1,2,3")
        assert dialect.delimiter == ","

    def test_tab_delimiter(self):
        dialect = sniff_csv_dialect("a\tb\tc\n1\t2\t3")
        assert dialect.delimiter == "\t"

    def test_fallback_on_unsniffable(self):
        # Should not raise; returns excel dialect
        dialect = sniff_csv_dialect("no delimiters here at all")
        assert dialect is not None


# ---------------------------------------------------------------------------
# decode_csv_content
# ---------------------------------------------------------------------------

class TestDecodeCsvContent:
    def test_utf8_sig(self, tmp_path: Path):
        path = tmp_path / "f.csv"
        path.write_bytes("归属地,城市\n广东,深圳\n".encode("utf-8-sig"))
        result = decode_csv_content(path)
        assert "归属地" in result

    def test_gbk_encoding(self, tmp_path: Path):
        path = tmp_path / "f.csv"
        path.write_bytes("归属地,城市\n广东,深圳\n".encode("gbk"))
        result = decode_csv_content(path)
        assert "归属地" in result


# ---------------------------------------------------------------------------
# read_xlsx_location_column
# ---------------------------------------------------------------------------

class TestReadXlsxLocationColumn:
    def test_reads_column(self, tmp_path: Path):
        path = _make_xlsx(tmp_path, [["归属地", "其他"], ["广东-深圳", "x"], ["北京-朝阳", "y"]])
        result = read_xlsx_location_column(path, "归属地")
        assert result == ["广东-深圳", "北京-朝阳"]

    def test_missing_column_raises(self, tmp_path: Path):
        path = _make_xlsx(tmp_path, [["城市", "人口"], ["深圳", 100]])
        with pytest.raises(ValueError, match="没有找到列"):
            read_xlsx_location_column(path, "归属地")

    def test_empty_file_raises(self, tmp_path: Path):
        path = _make_xlsx(tmp_path, [])
        with pytest.raises(ValueError, match="没有表头行"):
            read_xlsx_location_column(path, "归属地")


# ---------------------------------------------------------------------------
# read_csv_location_column
# ---------------------------------------------------------------------------

class TestReadCsvLocationColumn:
    def test_reads_column(self, tmp_path: Path):
        path = _make_csv(tmp_path, "归属地,其他\n广东-深圳,x\n北京-朝阳,y\n")
        result = read_csv_location_column(path, "归属地")
        assert result == ["广东-深圳", "北京-朝阳"]

    def test_missing_column_raises(self, tmp_path: Path):
        path = _make_csv(tmp_path, "城市,人口\n深圳,100\n")
        with pytest.raises(ValueError, match="没有找到列"):
            read_csv_location_column(path, "归属地")


# ---------------------------------------------------------------------------
# read_location_column (dispatcher)
# ---------------------------------------------------------------------------

class TestReadLocationColumn:
    def test_xlsx_dispatched(self, tmp_path: Path):
        path = _make_xlsx(tmp_path, [["归属地"], ["广东-深圳"]])
        result = read_location_column(path, "归属地")
        assert result == ["广东-深圳"]

    def test_csv_dispatched(self, tmp_path: Path):
        path = _make_csv(tmp_path, "归属地\n北京-朝阳\n")
        result = read_location_column(path, "归属地")
        assert result == ["北京-朝阳"]

    def test_unsupported_extension_raises(self, tmp_path: Path):
        path = tmp_path / "data.txt"
        path.write_text("hello")
        with pytest.raises(ValueError, match="不支持的文件类型"):
            read_location_column(path, "归属地")


# ---------------------------------------------------------------------------
# resolve_config_path
# ---------------------------------------------------------------------------

class TestResolveConfigPath:
    def test_explicit_existing_path(self, tmp_path: Path):
        cfg = _make_yaml(tmp_path, {"tiers": {"first": ["北京市"]}})
        result = resolve_config_path(cfg)
        assert result == cfg

    def test_explicit_missing_path_raises(self):
        with pytest.raises(FileNotFoundError):
            resolve_config_path("/nonexistent/path/tiers.yaml")

    def test_no_path_finds_default(self, tmp_path: Path, monkeypatch):
        cfg = tmp_path / "city_tiers.yaml"
        cfg.write_text(yaml.dump({"tiers": {"first": ["北京市"]}}, allow_unicode=True))
        monkeypatch.chdir(tmp_path)
        result = resolve_config_path()
        assert result == cfg


# ---------------------------------------------------------------------------
# load_city_tier_mapping
# ---------------------------------------------------------------------------

class TestLoadCityTierMapping:
    def test_loads_mapping(self, tmp_path: Path):
        cfg = _make_yaml(tmp_path, {"tiers": {"first": ["北京市", "上海市"]}})
        mapping = load_city_tier_mapping(cfg)
        assert mapping.get("北京") == "first"
        assert mapping.get("上海") == "first"

    def test_missing_tiers_key_raises(self, tmp_path: Path):
        cfg = tmp_path / "bad.yaml"
        cfg.write_text(yaml.dump({"cities": []}))
        with pytest.raises(ValueError, match="缺少 tiers"):
            load_city_tier_mapping(cfg)

    def test_unknown_tier_raises(self, tmp_path: Path):
        cfg = _make_yaml(tmp_path, {"tiers": {"super": ["北京市"]}})
        with pytest.raises(ValueError, match="未知城市分层"):
            load_city_tier_mapping(cfg)

    def test_duplicate_city_raises(self, tmp_path: Path):
        cfg = _make_yaml(
            tmp_path,
            {"tiers": {"first": ["北京市"], "new_first": ["北京市"]}},
        )
        with pytest.raises(ValueError, match="城市重复"):
            load_city_tier_mapping(cfg)

    def test_diqu_city_key_expansion(self, tmp_path: Path):
        cfg = _make_yaml(tmp_path, {"tiers": {"fourth": ["喀什地区"]}})
        mapping = load_city_tier_mapping(cfg)
        assert mapping.get("喀什") == "fourth"
        assert mapping.get("喀什地区") == "fourth"

    def test_none_city_list_skipped(self, tmp_path: Path):
        cfg = _make_yaml(tmp_path, {"tiers": {"first": None, "second": ["厦门市"]}})
        mapping = load_city_tier_mapping(cfg)
        assert mapping.get("厦门") == "second"


# ---------------------------------------------------------------------------
# write_location_details
# ---------------------------------------------------------------------------

class TestWriteLocationDetails:
    SAMPLE_DETAILS = [
        {
            "归属地": "广东-深圳",
            "城市": "深圳",
            "城市归一化": "深圳",
            "分层代码": "first",
            "分层": "一线",
        }
    ]

    def test_write_csv(self, tmp_path: Path):
        path = tmp_path / "out.csv"
        write_location_details(self.SAMPLE_DETAILS, path)
        content = path.read_text(encoding="utf-8-sig")
        assert "深圳" in content
        assert "一线" in content

    def test_write_xlsx(self, tmp_path: Path):
        path = tmp_path / "out.xlsx"
        write_location_details(self.SAMPLE_DETAILS, path)
        wb = lw(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "归属地"
        assert rows[1][1] == "深圳"

    def test_unsupported_extension_raises(self, tmp_path: Path):
        path = tmp_path / "out.json"
        with pytest.raises(ValueError, match="只支持"):
            write_location_details(self.SAMPLE_DETAILS, path)


# ---------------------------------------------------------------------------
# Integration: end-to-end flow using bundled city_tiers.yaml
# ---------------------------------------------------------------------------

class TestIntegration:
    """End-to-end tests using the real city_tiers.yaml config."""

    @pytest.fixture()
    def city_tier_mapping_from_config(self) -> dict[str, str]:
        config = Path(__file__).resolve().parent.parent / "city_tiers.yaml"
        return load_city_tier_mapping(config)

    def test_beijing_is_first(self, city_tier_mapping_from_config):
        # 直辖市格式: 省份-城市-区, 北京市的城市段也是"北京"
        assert match_location_tier("北京市-北京-朝阳区", city_tier_mapping_from_config) == "first"

    def test_chengdu_is_new_first(self, city_tier_mapping_from_config):
        assert match_location_tier("四川-成都-高新区", city_tier_mapping_from_config) == "new_first"

    def test_xiamen_is_second(self, city_tier_mapping_from_config):
        assert match_location_tier("福建-厦门-思明区", city_tier_mapping_from_config) == "second"

    def test_kashgar_is_fourth(self, city_tier_mapping_from_config):
        assert match_location_tier("新疆-喀什-喀什市", city_tier_mapping_from_config) == "fourth"

    def test_unknown_is_other(self, city_tier_mapping_from_config):
        assert match_location_tier("火星-某市-某区", city_tier_mapping_from_config) == "other"

    def test_calculate_stats_with_real_mapping(self, city_tier_mapping_from_config):
        locations = [
            "北京市-北京-朝阳",
            "四川-成都-高新",
            "新疆-喀什-喀什市",
            "未知省-未知市",
        ]
        stats = calculate_tier_stats(locations, city_tier_mapping_from_config)
        tier_map = {row["tier"]: row for row in stats}
        assert tier_map["first"]["count"] == 1
        assert tier_map["new_first"]["count"] == 1
        assert tier_map["fourth"]["count"] == 1
        assert tier_map["other"]["count"] == 1
