#!/usr/bin/env python3
"""City tier statistics helpers."""

from __future__ import annotations

import argparse
import csv
import io
import sys
import warnings
from collections import Counter
from pathlib import Path
from typing import Any, Final

import yaml
from openpyxl import Workbook, load_workbook


DEFAULT_CITY_TIERS_FILE: Final = "city_tiers.yaml"
LOCATION_COLUMN: Final = "归属地"
DETAIL_COLUMNS: Final = ("归属地", "城市", "城市归一化", "分层代码", "分层")
CSV_ENCODINGS: Final = (
    "utf-8-sig",
    "utf-16",
    "utf-16le",
    "utf-16be",
    "gb18030",
    "gbk",
    "big5",
    "cp950",
)


TIER_ORDER: Final = (
    "new_first",
    "first",
    "second",
    "third",
    "fourth",
    "other",
)


TIER_LABELS: Final = {
    "new_first": "新一线",
    "first": "一线",
    "second": "二线",
    "third": "三线",
    "fourth": "四线",
    "other": "其他",
}


def bundled_base_dir() -> Path | None:
    """Return PyInstaller's unpacked bundle directory when available."""
    bundle_dir = getattr(sys, "_MEIPASS", None)
    if bundle_dir:
        return Path(bundle_dir)
    return None


def default_config_candidates() -> list[Path]:
    """Build default config lookup paths."""
    candidates = [Path.cwd() / DEFAULT_CITY_TIERS_FILE]

    executable_dir = Path(sys.executable).resolve().parent
    script_dir = Path(__file__).resolve().parent
    for base_dir in (executable_dir, script_dir, bundled_base_dir()):
        if base_dir is None:
            continue
        candidate = base_dir / DEFAULT_CITY_TIERS_FILE
        if candidate not in candidates:
            candidates.append(candidate)

    return candidates


def resolve_config_path(config_path: str | Path | None = None) -> Path:
    """Resolve user config or the bundled/default city tiers config."""
    if config_path:
        path = Path(config_path)
        if not path.exists():
            raise FileNotFoundError(f"找不到城市分层配置文件：{path}")
        return path

    for path in default_config_candidates():
        if path.exists():
            return path

    searched = "\n".join(f"- {path}" for path in default_config_candidates())
    raise FileNotFoundError(f"找不到默认城市分层配置文件，已查找：\n{searched}")


def is_empty_value(value: Any) -> bool:
    """Return whether a raw spreadsheet value should be treated as empty."""
    if value is None:
        return True
    try:
        return bool(value != value)
    except TypeError:
        return False


def read_xlsx_location_column(path: Path, column_name: str) -> list[Any]:
    """Read the location column from an XLSX file."""
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            raise ValueError("文件没有表头行") from None

        if column_name not in headers:
            raise ValueError(f"文件中没有找到列：{column_name}")

        column_index = headers.index(column_name)
        locations: list[Any] = []
        for row in rows:
            value = row[column_index] if column_index < len(row) else None
            locations.append(value)
        return locations
    finally:
        workbook.close()


def decode_csv_content(path: Path) -> str:
    """Decode CSV bytes with common encodings used by Excel and Chinese systems."""
    content = path.read_bytes()
    last_error: UnicodeDecodeError | None = None
    for encoding in CSV_ENCODINGS:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError as error:
            last_error = error

    try:
        return content.decode("gb18030", errors="replace")
    except UnicodeDecodeError:
        raise ValueError(f"无法识别 CSV 文件编码，已尝试：{', '.join(CSV_ENCODINGS)}") from last_error


def sniff_csv_dialect(sample: str) -> csv.Dialect:
    """Detect CSV delimiter, falling back to Excel CSV style."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;，")
    except csv.Error:
        return csv.excel


def clean_csv_header(header: str | None) -> str:
    """Normalize CSV header text."""
    if header is None:
        return ""
    return header.strip().lstrip("\ufeff\ufffe\ufffd")


def read_csv_location_column(path: Path, column_name: str) -> list[Any]:
    """Read the location column from a CSV file."""
    content = decode_csv_content(path)
    reader = csv.DictReader(
        io.StringIO(content),
        dialect=sniff_csv_dialect(content[:4096]),
    )
    if reader.fieldnames is None:
        raise ValueError("文件没有表头行")

    header_map = {clean_csv_header(header): header for header in reader.fieldnames}
    if column_name not in header_map:
        found_columns = ", ".join(clean_csv_header(header) for header in reader.fieldnames)
        raise ValueError(f"文件中没有找到列：{column_name}；当前列：{found_columns}")

    raw_column_name = header_map[column_name]
    locations: list[Any] = []
    for row in reader:
        locations.append(row.get(raw_column_name))
    return locations


def read_location_column(file_path: str | Path, column_name: str = LOCATION_COLUMN) -> list[Any]:
    """Read the location column from an Excel or CSV file."""
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return read_xlsx_location_column(path, column_name)
    elif suffix == ".csv":
        return read_csv_location_column(path, column_name)
    else:
        raise ValueError(f"不支持的文件类型：{suffix or '无扩展名'}")


def extract_city_name(raw_location: Any) -> str:
    """Extract city from values like '重庆-重庆-九龙坡'."""
    if is_empty_value(raw_location):
        return ""

    location = str(raw_location).strip()
    if not location:
        return ""

    for separator in ("－", "—", "–"):
        location = location.replace(separator, "-")

    parts = [part.strip() for part in location.split("-") if part.strip()]
    if len(parts) >= 2:
        return parts[1]
    return location


def normalize_city_name(city_name: Any) -> str:
    """Normalize city names before matching."""
    if is_empty_value(city_name):
        return ""

    city = str(city_name).strip()
    if city.endswith("市"):
        city = city[:-1]
    return city


def city_match_keys(city_name: Any) -> set[str]:
    """Build practical match keys for city names in config and input."""
    city = normalize_city_name(city_name)
    if not city:
        return set()

    keys = {city}
    for suffix in ("地区", "盟"):
        if city.endswith(suffix):
            keys.add(city[: -len(suffix)])

    if city.endswith("自治州"):
        keys.add(city[: -len("自治州")])

    return keys


def load_city_tier_mapping(config_path: str | Path | None = None) -> dict[str, str]:
    """Load city-to-tier mapping from YAML."""
    path = resolve_config_path(config_path)
    with path.open("r", encoding="utf-8") as file:
        config = yaml.safe_load(file) or {}

    tiers = config.get("tiers")
    if not isinstance(tiers, dict):
        raise ValueError("YAML 配置缺少 tiers 字段")

    city_to_tier: dict[str, str] = {}
    valid_tiers = set(TIER_ORDER) - {"other"}

    for tier, cities in tiers.items():
        if tier not in valid_tiers:
            raise ValueError(f"YAML 配置中存在未知城市分层：{tier}")
        if cities is None:
            continue
        if not isinstance(cities, list):
            raise ValueError(f"YAML 配置中 {tier} 必须是城市列表")

        for city in cities:
            for key in city_match_keys(city):
                existing_tier = city_to_tier.get(key)
                if existing_tier and existing_tier != tier:
                    raise ValueError(f"城市重复出现在多个分层：{city}")
                city_to_tier[key] = tier

    return city_to_tier


def match_city_tier(city_name: Any, city_to_tier: dict[str, str]) -> str:
    """Match a city name to its tier, defaulting to other."""
    for key in city_match_keys(city_name):
        tier = city_to_tier.get(key)
        if tier:
            return tier
    return "other"


def match_location_tier(raw_location: Any, city_to_tier: dict[str, str]) -> str:
    """Extract city from a raw location value and match it to a tier."""
    city_name = extract_city_name(raw_location)
    return match_city_tier(city_name, city_to_tier)


def calculate_tier_stats(
    locations: list[Any],
    city_to_tier: dict[str, str],
) -> list[dict[str, int | float | str]]:
    """Count city tiers and percentages from location values."""
    tiers = [match_location_tier(value, city_to_tier) for value in locations]
    counts = Counter(tiers)
    total = len(locations)

    results: list[dict[str, int | float | str]] = []
    for tier in TIER_ORDER:
        count = counts.get(tier, 0)
        percentage = count / total if total else 0
        results.append(
            {
                "tier": tier,
                "label": TIER_LABELS[tier],
                "count": count,
                "percentage": percentage,
            }
        )
    return results


def build_location_details(
    locations: list[Any],
    city_to_tier: dict[str, str],
) -> list[dict[str, Any]]:
    """Build row-level location matching details."""
    details: list[dict[str, Any]] = []
    for raw_location in locations:
        city = extract_city_name(raw_location)
        normalized_city = normalize_city_name(city)
        tier = match_city_tier(city, city_to_tier)
        details.append(
            {
                "归属地": raw_location,
                "城市": city,
                "城市归一化": normalized_city,
                "分层代码": tier,
                "分层": TIER_LABELS[tier],
            }
        )
    return details


def write_location_details(details: list[dict[str, Any]], output_path: str | Path) -> None:
    """Write row-level location matching details to a CSV or XLSX file."""
    path = Path(output_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(DETAIL_COLUMNS)
        for row in details:
            sheet.append([row.get(column, "") for column in DETAIL_COLUMNS])
        workbook.save(path)
    elif suffix == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=DETAIL_COLUMNS)
            writer.writeheader()
            writer.writerows(details)
    else:
        raise ValueError(f"明细文件只支持 .xlsx、.csv：{suffix or '无扩展名'}")


def output_tier_stats(stats: list[dict[str, int | float | str]]) -> None:
    """Print city tier statistics."""
    total = sum(int(row["count"]) for row in stats)

    print(f"总数: {total}")

    for row in stats:
        label = str(row["label"])
        count = int(row["count"])
        percentage = float(row["percentage"])
        print(f"{label}: {count}, {percentage:.2%}")


def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="统计归属地城市分层数量和占比")
    parser.add_argument("file", help="输入文件，支持 .xlsx、.csv")
    parser.add_argument(
        "-c",
        "--config",
        help=f"自定义城市分层 YAML 配置文件；不传则自动查找 {DEFAULT_CITY_TIERS_FILE}",
    )
    parser.add_argument(
        "--column",
        default=LOCATION_COLUMN,
        help=f"归属地列名，默认：{LOCATION_COLUMN}",
    )
    parser.add_argument(
        "--detail-output",
        help="导出逐行匹配明细，支持 .xlsx、.csv；默认不导出",
    )
    return parser.parse_args()


def main() -> None:
    """Run command line entrypoint."""
    args = parse_args()
    city_to_tier = load_city_tier_mapping(args.config)
    locations = read_location_column(args.file, args.column)
    stats = calculate_tier_stats(locations, city_to_tier)
    output_tier_stats(stats)

    if args.detail_output:
        details = build_location_details(locations, city_to_tier)
        write_location_details(details, args.detail_output)
        print(f"明细已导出: {args.detail_output}")


if __name__ == "__main__":
    main()
